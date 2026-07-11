#!/usr/bin/env python3
"""
Porsche Sales Sanitizer Agent - schema_new.md version

Reads a Porsche sales Excel file, applies the sanitization rules from schema_new.md,
and generates a treated Excel file with sanitized columns inserted immediately after
source columns.

Main correction versus the previous version:
- Invalid dates caused by an excessive day number are corrected by rolling the
  excess days into the following month. Examples:
    2024-02-30 -> 2024-03-01
    April 31st, 2024 -> 2024-05-01
    2027-06-40 -> 2027-07-10

Usage in Anaconda Prompt:
    python porsche_sanitizer_agent_schema_new.py --input porsche_data_base.xlsx --schema schema_new.md --output porsche_data_base_sanitize.xlsx

Usage in Jupyter Notebook:
    !python porsche_sanitizer_agent_schema_new.py --input porsche_data_base.xlsx --schema schema_new.md --output porsche_data_base_sanitize.xlsx

Dependencies:
    pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
import calendar
import math
import re
import string
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("Missing dependency: pandas. Install with: pip install pandas openpyxl") from exc


REQUIRED_COLUMNS = [
    "sale_id",
    "sale_date",
    "customer_name",
    "porsche_model",
    "model_year",
    "sale_price",
    "vehicle_mileage",
    "payment_method",
    "city",
    "state",
    "salesperson",
    "delivery_status",
]

CANONICAL_MODELS = [
    "911 Carrera", "911 Carrera S", "911 Carrera GTS", "911 Turbo", "911 Turbo S",
    "911 GT3", "911 GT3 RS", "911 Dakar", "911 Targa 4", "911 Targa 4S",
    "718 Cayman", "718 Cayman S", "718 Cayman GT4 RS", "718 Boxster",
    "718 Boxster GTS", "718 Spyder RS", "Cayenne", "Cayenne S", "Cayenne Coupe",
    "Cayenne E-Hybrid", "Cayenne Turbo", "Cayenne Turbo GT", "Macan", "Macan S",
    "Macan T", "Macan GTS", "Macan Electric", "Panamera", "Panamera 4",
    "Panamera 4S", "Panamera Turbo", "Panamera Turbo S", "Panamera 4 E-Hybrid",
    "Taycan", "Taycan 4S", "Taycan GTS", "Taycan Turbo", "Taycan Turbo S",
    "Taycan Cross Turismo",
]

UNITS = {
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
    "seven": 7, "eight": 8, "nine": 9, "ten": 10, "eleven": 11, "twelve": 12,
    "thirteen": 13, "fourteen": 14, "fifteen": 15, "sixteen": 16, "seventeen": 17,
    "eighteen": 18, "nineteen": 19,
}
TENS = {
    "twenty": 20, "thirty": 30, "forty": 40, "fourty": 40, "fifty": 50,
    "sixty": 60, "seventy": 70, "eighty": 80, "ninety": 90,
}
SCALE = {"hundred": 100, "thousand": 1000, "million": 1000000}
FILLER = {
    "and", "usd", "dollar", "dollars", "mile", "miles", "mi", "km", "kms",
    "kilometer", "kilometers", "kilometre", "kilometres", "car", "new", "cars",
}

MONTHS = {name.lower(): i for i, name in enumerate(calendar.month_name) if name}
MONTHS.update({name.lower(): i for i, name in enumerate(calendar.month_abbr) if name})


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip() == ""


def norm_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


MODEL_MAP = {norm_key(model): model for model in CANONICAL_MODELS}


def words_to_number(text: Any) -> int | None:
    if text is None:
        return None

    normalized = str(text).lower()
    normalized = re.sub(r"[^a-z\s-]", " ", normalized).replace("-", " ")
    tokens = [token for token in normalized.split() if token not in FILLER]

    if not tokens:
        return None

    total = 0
    current = 0
    matched = False

    for token in tokens:
        if token in UNITS:
            current += UNITS[token]
            matched = True
        elif token in TENS:
            current += TENS[token]
            matched = True
        elif token == "hundred":
            current = (current if current else 1) * 100
            matched = True
        elif token in ("thousand", "million"):
            total += (current if current else 1) * SCALE[token]
            current = 0
            matched = True
        else:
            return None

    return total + current if matched else None


def round_half_up(number: float | int) -> int:
    return int(Decimal(str(number)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def parse_localized_number(text: Any) -> float | None:
    if text is None:
        return None

    if isinstance(text, (int, float)) and not isinstance(text, bool):
        if isinstance(text, float) and math.isnan(text):
            return None
        return float(text)

    match = re.search(r"-?\d[\d,\.]*", str(text).strip())
    if not match:
        return None

    value = match.group(0)

    if "," not in value and "." not in value:
        try:
            return float(value)
        except ValueError:
            return None

    if "," in value and "." in value:
        last_comma = value.rfind(",")
        last_dot = value.rfind(".")
        decimal_separator = "," if last_comma > last_dot else "."
        fraction_length = len(value.split(decimal_separator)[-1])

        if fraction_length in (1, 2):
            thousands_separator = "." if decimal_separator == "," else ","
            cleaned = value.replace(thousands_separator, "").replace(decimal_separator, ".")
        else:
            cleaned = value.replace(",", "").replace(".", "")

    elif "," in value:
        parts = value.split(",")
        if len(parts) > 2:
            cleaned = "".join(parts)
        else:
            before, after = parts
            if len(after) == 3:
                cleaned = before + after
            elif len(after) in (1, 2):
                cleaned = before + "." + after
            else:
                cleaned = before + after

    else:
        parts = value.split(".")
        if len(parts) > 2:
            cleaned = "".join(parts)
        else:
            before, after = parts
            if len(after) == 3:
                cleaned = before + after
            elif len(after) in (1, 2):
                cleaned = before + "." + after
            else:
                cleaned = before + after

    try:
        return float(cleaned)
    except ValueError:
        return None


def two_digit_year_to_four(year: str | int) -> int:
    year_int = int(year)
    if year_int < 100:
        return 2000 + year_int
    return year_int


def build_date(year: int, month: int, day: int, correct_overflow_day: bool = True) -> date | None:
    """Build a date, optionally correcting excessive day numbers per schema_new.md."""
    if not (1 <= month <= 12) or day < 1:
        return None

    try:
        return date(year, month, day)
    except ValueError:
        if not correct_overflow_day:
            return None

        last_day = calendar.monthrange(year, month)[1]
        if day > last_day:
            excess_days = day - last_day
            return date(year, month, last_day) + timedelta(days=excess_days)
        return None


def _parse_date_components(value: Any, correct_overflow_day: bool) -> date | None:
    if is_blank(value):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        # Excel serial date support.
        if 1 <= float(value) <= 60000:
            try:
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
            except Exception:
                return None
        return None

    raw = str(value).strip()
    if not raw:
        return None

    # Sometimes Excel-like datetimes are passed as text.
    raw = re.sub(r"\s+00:00:00$", "", raw)
    cleaned = re.sub(r"(\d{1,2})(st|nd|rd|th)\b", r"\1", raw, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", " ", cleaned.replace(",", ", ").strip())
    cleaned = re.sub(r"\s+,", ",", cleaned)

    # Month-name formats: Month DD, YYYY / Mon DD YYYY.
    month_match = re.fullmatch(r"([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})", cleaned)
    if month_match:
        month_name, day_text, year_text = month_match.groups()
        month = MONTHS.get(month_name.lower())
        if month is None:
            return None
        return build_date(int(year_text), month, int(day_text), correct_overflow_day)

    # Year-first numeric formats: YYYY-MM-DD, YYYY/MM/DD, YYYY/DD/MM, YYYY.MM.DD.
    year_first = re.fullmatch(r"(\d{4})([-/.])(\d{1,2})\2(\d{1,2})", cleaned)
    if year_first:
        year_text, sep, second_text, third_text = year_first.groups()
        year = int(year_text)
        second = int(second_text)
        third = int(third_text)

        if sep == "/":
            # Disambiguate between YYYY/MM/DD and YYYY/DD/MM.
            if second > 12 and third <= 12:
                day, month = second, third
            else:
                month, day = second, third
        else:
            # Hyphen and dot are treated as YYYY-MM-DD / YYYY.MM.DD.
            month, day = second, third

        return build_date(year, month, day, correct_overflow_day)

    # US numeric formats: MM/DD/YYYY, MM/DD/YY, MM-DD-YY.
    us_match = re.fullmatch(r"(\d{1,2})([-/])(\d{1,2})\2(\d{2}|\d{4})", cleaned)
    if us_match:
        month_text, _sep, day_text, year_text = us_match.groups()
        return build_date(two_digit_year_to_four(year_text), int(month_text), int(day_text), correct_overflow_day)

    return None


def sanitize_date(value: Any) -> str:
    parsed = _parse_date_components(value, correct_overflow_day=True)
    return parsed.strftime("%Y-%m-%d") if parsed else "INVALID"


def strict_sanitize_date(value: Any) -> str:
    parsed = _parse_date_components(value, correct_overflow_day=False)
    return parsed.strftime("%Y-%m-%d") if parsed else "INVALID"


def date_corrected_by_overflow_rule(value: Any) -> bool:
    strict = strict_sanitize_date(value)
    corrected = sanitize_date(value)
    return strict == "INVALID" and corrected != "INVALID"


def smart_title_model(value: Any) -> str:
    raw = re.sub(r"\s+", " ", str(value).strip())
    if not raw:
        return "INVALID"

    titled = string.capwords(raw.lower())
    replacements = {
        "Gt3": "GT3",
        "Gt4": "GT4",
        "Gts": "GTS",
        "Rs": "RS",
        "E-hybrid": "E-Hybrid",
        "4s": "4S",
    }

    for source, target in replacements.items():
        titled = re.sub(rf"\b{re.escape(source)}\b", target, titled)

    return titled


def sanitize_model(value: Any) -> str:
    if is_blank(value):
        return "INVALID"

    return MODEL_MAP.get(norm_key(value), smart_title_model(value))


def sanitize_model_year(value: Any) -> str:
    if is_blank(value):
        return "INVALID"

    year = None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if not (isinstance(value, float) and math.isnan(value)) and float(value).is_integer():
            year = int(value)
    else:
        raw = str(value).strip().lower()

        if re.fullmatch(r"\d{4}", raw):
            year = int(raw)
        elif re.fullmatch(r"\d{2}[-\s]\d{2}", raw) or re.fullmatch(r"\d{2}\D+\d{2}", raw):
            year = int(re.sub(r"\D", "", raw))
        else:
            tokens = re.sub(r"[^a-z\s-]", " ", raw).replace("-", " ").split()

            if len(tokens) >= 2 and tokens[0] == "twenty" and tokens[1] in TENS:
                parsed = words_to_number(" ".join(tokens[1:]))
                if parsed is not None:
                    year = 2000 + parsed

            if year is None:
                parsed = words_to_number(raw)
                if parsed is not None:
                    year = int(parsed)

    if year is None or not (1990 <= year <= 2035):
        return "INVALID"

    return str(year)


def sanitize_price(value: Any) -> str:
    if is_blank(value):
        return "INVALID"

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return "INVALID"
        return f"{float(value):.2f}"

    raw = str(value).strip().lower()

    if not re.search(r"\d", raw):
        parsed = words_to_number(raw)
        return "INVALID" if parsed is None else f"{float(parsed):.2f}"

    multiplier = 1000 if re.search(r"\d[\d,\.]*\s*k\b", raw) else 1
    parsed = parse_localized_number(raw)

    return "INVALID" if parsed is None else f"{parsed * multiplier:.2f}"


def sanitize_mileage(value: Any) -> str:
    if is_blank(value):
        return "INVALID"

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return "INVALID"
        return str(round_half_up(value))

    raw = str(value).strip().lower()

    if re.search(r"\b(new|new car|zero)\b", raw) or re.fullmatch(r"\s*0\s*(mi|miles)?\.?\s*", raw):
        return "0"

    is_km = bool(re.search(r"\bkm\b|\bkms\b|kilometers?|kilometres?", raw))
    parsed = parse_localized_number(raw)

    if parsed is None:
        parsed = words_to_number(raw)

    if parsed is None:
        return "INVALID"

    miles = parsed * 0.621371 if is_km else parsed
    return str(round_half_up(miles))


PAYMENT_MAP: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\bach\b"), "ACH Payment"),
    (re.compile(r"\bcrypto\b"), "Crypto Payment"),
    (re.compile(r"\bcredit\b|\bcreditcard\b"), "Credit Card"),
    (re.compile(r"\bdebit\b"), "Debit Card"),
    (re.compile(r"\b(bank\s*transfer|bank-transfer|bank_transfer)\b"), "Bank Transfer"),
    (re.compile(r"\bwire\b|\bwiretransfer\b"), "Wire Transfer"),
    (re.compile(r"\bfinanc"), "Financing"),
    (re.compile(r"\bleas"), "Lease"),
    (re.compile(r"\bcash\b"), "Cash"),
]


def sanitize_payment(value: Any) -> str:
    if is_blank(value):
        return "INVALID"

    key = norm_key(value)
    compact = re.sub(r"\s+", "", key)
    test_value = f"{key} {compact}"

    for pattern, label in PAYMENT_MAP:
        if pattern.search(test_value):
            return label

    return string.capwords(str(value).strip().replace("_", " ").replace("-", " "))


def sanitize_city(value: Any) -> str:
    if is_blank(value):
        return "INVALID"

    raw = re.sub(r"\s+", " ", str(value).strip())
    return "INVALID" if not raw else string.capwords(raw.lower())


US_STATES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas", "CA": "California",
    "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware", "FL": "Florida", "GA": "Georgia",
    "HI": "Hawaii", "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VT": "Vermont",
    "VA": "Virginia", "WA": "Washington", "WV": "West Virginia", "WI": "Wisconsin",
    "WY": "Wyoming", "DC": "District of Columbia",
}
STATE_BY_NAME = {state_name.lower(): state_code for state_code, state_name in US_STATES.items()}


def sanitize_state(value: Any) -> str:
    if is_blank(value):
        return "INVALID"

    raw = re.sub(r"\s+", " ", str(value).strip())
    upper = raw.upper()

    if upper in US_STATES:
        return upper

    return STATE_BY_NAME.get(raw.lower(), "INVALID")


DELIVERY_MAP = {
    "deliverd": "Delivered",
    "delivered": "Delivered",
    "pending": "Pending",
    "in transit": "In Transit",
    "cancelled": "Cancelled",
    "canceled": "Cancelled",
    "awaiting delivery": "Awaiting Delivery",
    "awaiting pickup": "Awaiting Pickup",
    "pending approval": "Pending Approval",
    "pending review": "Pending Review",
    "shipped": "Shipped",
    "awaiting review": "Awaiting Review",
}


def sanitize_delivery(value: Any) -> str:
    if is_blank(value):
        return "INVALID"

    key = norm_key(value)
    return DELIVERY_MAP.get(key, string.capwords(key))


SANITIZERS: dict[str, tuple[str, Callable[[Any], str]]] = {
    "sale_date": ("SaleDateSanitized", sanitize_date),
    "porsche_model": ("PorscheModelSanitized", sanitize_model),
    "model_year": ("ModelYearSanitized", sanitize_model_year),
    "sale_price": ("SalesPriceSanitized", sanitize_price),
    "vehicle_mileage": ("VehicleMileageSanitized", sanitize_mileage),
    "payment_method": ("PayMethodSanitized", sanitize_payment),
    "city": ("CitySanitized", sanitize_city),
    "state": ("StateSanitized", sanitize_state),
    "delivery_status": ("DeliveryStatusSanitized", sanitize_delivery),
}


def sanitize_dataframe(df: "pd.DataFrame") -> tuple["pd.DataFrame", "pd.DataFrame", "pd.DataFrame", int]:
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in df.columns]

    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    output = pd.DataFrame(index=df.index)

    date_overflow_corrections = 0

    for column in df.columns:
        output[column] = df[column]

        if column in SANITIZERS:
            sanitized_column, sanitizer = SANITIZERS[column]
            output[sanitized_column] = df[column].apply(lambda value: sanitizer(value) or "INVALID").replace("", "INVALID")

            if column == "sale_date":
                date_overflow_corrections = int(df[column].apply(date_corrected_by_overflow_rule).sum())

    sanitized_columns = [SANITIZERS[column][0] for column in df.columns if column in SANITIZERS]

    quality_rows = []
    for column in sanitized_columns:
        quality_rows.append(
            {
                "Sanitized Column": column,
                "INVALID Count": int((output[column].astype(str).str.strip() == "INVALID").sum()),
                "Blank Count": int(output[column].isna().sum() + (output[column].astype(str).str.strip() == "").sum()),
            }
        )

    position_rows = []
    output_columns = list(output.columns)
    for source_column, (sanitized_column, _) in SANITIZERS.items():
        source_position = output_columns.index(source_column)
        sanitized_position = output_columns.index(sanitized_column)
        position_rows.append(
            {
                "Source Column": source_column,
                "Sanitized Column": sanitized_column,
                "Immediately After Source": sanitized_position == source_position + 1,
            }
        )

    return output, pd.DataFrame(quality_rows), pd.DataFrame(position_rows), date_overflow_corrections


def write_output_excel(
    sanitized_df: "pd.DataFrame",
    quality_df: "pd.DataFrame",
    position_df: "pd.DataFrame",
    date_overflow_corrections: int,
    output_path: Path,
    source_file: Path,
    schema_file: Path,
    sheet_name: str,
) -> None:
    summary_df = pd.DataFrame(
        [
            {"Metric": "Source workbook", "Value": source_file.name},
            {"Metric": "Schema file", "Value": schema_file.name},
            {"Metric": "Source sheet", "Value": sheet_name},
            {"Metric": "Source rows processed", "Value": len(sanitized_df)},
            {"Metric": "Output columns", "Value": len(sanitized_df.columns)},
            {"Metric": "Sanitized columns added", "Value": len(quality_df)},
            {"Metric": "Date overflow corrections", "Value": date_overflow_corrections},
            {
                "Metric": "Sanitized columns immediately after source",
                "Value": "PASS" if bool(position_df["Immediately After Source"].all()) else "FAIL",
            },
            {"Metric": "Sanitized blanks introduced", "Value": int(quality_df["Blank Count"].sum())},
        ]
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        sanitized_df.to_excel(writer, index=False, sheet_name="Sanitized_Data")
        summary_df.to_excel(writer, index=False, sheet_name="Quality_Checks", startrow=0)
        quality_df.to_excel(writer, index=False, sheet_name="Quality_Checks", startrow=len(summary_df) + 3)
        position_df.to_excel(
            writer,
            index=False,
            sheet_name="Quality_Checks",
            startrow=len(summary_df) + len(quality_df) + 6,
        )

        workbook = writer.book

        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.worksheet.table import Table, TableStyleInfo

            data_ws = workbook["Sanitized_Data"]
            quality_ws = workbook["Quality_Checks"]

            header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in data_ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            data_ws.freeze_panes = "A2"

            for column_cells in data_ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                data_ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 30)

            data_table_ref = f"A1:{data_ws.cell(row=data_ws.max_row, column=data_ws.max_column).coordinate}"
            table = Table(displayName="PorscheSanitizedTable", ref=data_table_ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            data_ws.add_table(table)

            # Style the three Quality_Checks sections.
            section_header_rows = [1, len(summary_df) + 4, len(summary_df) + len(quality_df) + 7]
            for row_num in section_header_rows:
                for cell in quality_ws[row_num]:
                    if cell.value is not None:
                        cell.fill = header_fill
                        cell.font = header_font

            for column_cells in quality_ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                quality_ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 42)

            quality_ws.freeze_panes = "A2"

        except Exception:
            # Formatting is helpful but not required for the data treatment itself.
            pass


def main() -> None:
    parser = argparse.ArgumentParser(description="Sanitize Porsche sales Excel data according to schema_new.md.")
    parser.add_argument("--input", required=True, type=Path, help="Path to source .xlsx file.")
    parser.add_argument("--schema", required=True, type=Path, help="Path to schema_new.md.")
    parser.add_argument("--output", default=Path("porsche_data_base_sanitize.xlsx"), type=Path, help="Output .xlsx path.")
    parser.add_argument("--sheet", default=0, help="Sheet name or zero-based sheet index. Default: 0.")
    args = parser.parse_args()

    if not args.input.exists():
        raise FileNotFoundError(f"Input file not found: {args.input}")

    if not args.schema.exists():
        raise FileNotFoundError(f"Schema file not found: {args.schema}")

    # Read the schema for traceability. The executable rules are encoded in the functions above.
    _schema_text = args.schema.read_text(encoding="utf-8")

    sheet_arg: str | int
    try:
        sheet_arg = int(args.sheet)
    except ValueError:
        sheet_arg = args.sheet

    source_df = pd.read_excel(args.input, sheet_name=sheet_arg, dtype=object)
    sheet_name = str(args.sheet)

    sanitized_df, quality_df, position_df, date_overflow_corrections = sanitize_dataframe(source_df)
    write_output_excel(
        sanitized_df,
        quality_df,
        position_df,
        date_overflow_corrections,
        args.output,
        args.input,
        args.schema,
        sheet_name,
    )

    print(f"Saved sanitized file: {args.output}")
    print(f"Rows processed: {len(sanitized_df)}")
    print(f"Sanitized columns added: {len(quality_df)}")
    print(f"Date overflow corrections: {date_overflow_corrections}")
    print(f"Blank sanitized values introduced: {int(quality_df['Blank Count'].sum())}")


if __name__ == "__main__":
    main()
